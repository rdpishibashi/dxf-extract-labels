"""create_excel_output / create_region_excel_output のスモークテスト。

Excel 生成関数が最小の入力で例外なく完走し、
正しいシート構成の Excel を返すことを検証する。
ロジックの詳細（書式・列幅）ではなく「動作する・壊れていない」を確認する。
"""
import sys
import os
from io import BytesIO

import openpyxl
import pytest

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, PROJECT_ROOT)

from utils.excel_output import (  # noqa: E402
    create_excel_output, create_ref_designator_excel_output, create_region_excel_output,
)


# ---------------------------------------------------------------------------
# ヘルパー
# ---------------------------------------------------------------------------

def _load_wb(data: bytes):
    return openpyxl.load_workbook(BytesIO(data))


# ---------------------------------------------------------------------------
# create_excel_output
# ---------------------------------------------------------------------------

def test_create_excel_output_returns_bytes():
    results = {
        '/tmp/foo.dxf': (['R10', 'CN3', 'R10'], {
            'filename': 'foo.dxf',
            'final_count': 3,
            'main_drawing_number': 'EE0001-001-01A',
            'source_drawing_number': '',
            'title': 'TEST',
            'subtitle': '',
        }),
    }
    data = create_excel_output(results, 'asc')
    assert isinstance(data, bytes)
    assert len(data) > 0


def test_create_excel_output_sheet_names():
    results = {
        '/tmp/circuit.dxf': (['R10', 'CN3'], {
            'filename': 'circuit.dxf',
            'final_count': 2,
            'main_drawing_number': '',
            'source_drawing_number': '',
            'title': '',
            'subtitle': '',
        }),
    }
    wb = _load_wb(create_excel_output(results, 'asc'))
    assert 'Summary' in wb.sheetnames
    assert 'Total' in wb.sheetnames
    assert 'circuit' in wb.sheetnames   # 拡張子なし


def test_create_excel_output_sort_option_desc():
    """sort_option='desc' でも例外なく完走する（フィルタ・妥当性チェックは v1.6.0 で削除済み）"""
    results = {
        '/tmp/x.dxf': (['R10', 'CHAMBER', 'CN3'], {
            'filename': 'x.dxf',
            'final_count': 3,
            'main_drawing_number': '',
            'source_drawing_number': '',
            'title': '',
            'subtitle': '',
        }),
    }
    data = create_excel_output(results, 'desc')
    wb = _load_wb(data)
    assert 'Summary' in wb.sheetnames


def test_create_excel_output_empty_results():
    """ファイルが0件でも例外なく返る"""
    data = create_excel_output({}, 'asc')
    wb = _load_wb(data)
    assert 'Summary' in wb.sheetnames


# ---------------------------------------------------------------------------
# create_ref_designator_excel_output
# ---------------------------------------------------------------------------

def _make_ref_designator_results():
    return {
        'fileA.dxf': {
            'rows': [{'ラベル': 'CN1', '個数': 2}, {'ラベル': 'R10', '個数': 1}],
            'total_in_frame': 10,
            'unclassified_count': 3,
            'main_drawing_number': 'EE1234-500-01A',
            'source_drawing_number': 'EE0001-500-01A',
            'title': 'TITLE A',
            'subtitle': 'SUB A',
        },
        'fileB.dxf': {
            'rows': [],
            'total_in_frame': 5,
            'unclassified_count': 0,
            'main_drawing_number': None,
            'source_drawing_number': None,
            'title': None,
            'subtitle': None,
        },
    }


def test_create_ref_designator_excel_output_returns_bytes():
    data = create_ref_designator_excel_output(_make_ref_designator_results(), 'asc')
    assert isinstance(data, bytes)
    assert len(data) > 0


def test_create_ref_designator_excel_output_sheet_and_values():
    """機器符号（候補）0件のファイル（fileB）はシートを作らない
    （`rows` が空のファイルシートはスキップする従来挙動を固定する）。"""
    wb = _load_wb(create_ref_designator_excel_output(_make_ref_designator_results(), 'asc'))
    assert wb.sheetnames == ['Summary', 'Total', 'fileA']

    summary = wb['Summary']
    header = [c.value for c in summary[1]]
    assert header == ['ファイル名', '図面枠内ラベル数', '機器符号（候補）数',
                      '未確定ラベル数（未採用）', '図番', '流用元図番', 'タイトル', 'サブタイトル']
    row_a = [c.value for c in summary[2]]
    assert row_a == ['fileA.dxf', 10, 3, 3, 'EE1234-500-01A', 'EE0001-500-01A', 'TITLE A', 'SUB A']

    total_rows = [(c[0].value, c[1].value) for c in wb['Total'].iter_rows(min_row=2)]
    assert total_rows == [('CN1', 2), ('R10', 1)]


def test_create_ref_designator_excel_output_sort_desc():
    wb = _load_wb(create_ref_designator_excel_output(_make_ref_designator_results(), 'desc'))
    labels = [row[0].value for row in wb['fileA'].iter_rows(min_row=2)]
    assert labels == ['R10', 'CN1']


# ---------------------------------------------------------------------------
# create_region_excel_output
# ---------------------------------------------------------------------------

def _make_region_results(fname='circuit.dxf'):
    return {
        fname: {
            'rows': [
                {'ラベル': 'R10', '個数': 2, '領域': 'RACK1'},
                {'ラベル': 'CN3', '個数': 1, '領域': ''},
            ],
            'named': [
                {'polygon': [], 'name': 'RACK1', 'id': 0,
                 'frame': 0, 'area_pct': 25.0, 'label_count': 2},
            ],
            'frames': 1,
            'regions_detected': 1,
            'regions_named': 1,
            'total_in_frame': 3,
            'filtered_count': 0,
            'final_count': 3,
            'in_region_count': 2,
        }
    }


def test_create_region_excel_output_returns_bytes():
    data = create_region_excel_output(_make_region_results())
    assert isinstance(data, bytes)
    assert len(data) > 0


def test_create_region_excel_output_sheet_names():
    wb = _load_wb(create_region_excel_output(_make_region_results()))
    assert 'Summary' in wb.sheetnames
    assert '領域一覧' in wb.sheetnames
    assert 'circuit' in wb.sheetnames


def test_create_region_excel_output_summary_row_count():
    """Summary シートの行数 = ヘッダー1行 + ファイル数"""
    results = _make_region_results()
    wb = _load_wb(create_region_excel_output(results))
    ws = wb['Summary']
    assert ws.max_row == 2   # ヘッダー + 1ファイル


def test_create_region_excel_output_empty():
    """ファイルなしでも例外なく返る"""
    data = create_region_excel_output({})
    wb = _load_wb(data)
    assert 'Summary' in wb.sheetnames


# ---------------------------------------------------------------------------
# 半角正規化（出力ファイルのラベル・領域名はすべて半角で集計・記録する）
# ---------------------------------------------------------------------------

def test_normalize_width_basic():
    from utils.common_utils import normalize_width
    assert normalize_width('ＳＹＳＴＥＭ　Ｉ／Ｆ　ＢＯＸ') == 'SYSTEM I/F BOX'
    assert normalize_width('ＣＮ１') == 'CN1'
    assert normalize_width('CN1') == 'CN1'
    assert normalize_width('ラック１') == 'ラック1'   # かなは不変・数字のみ半角化
    assert normalize_width('') == ''


def test_create_excel_output_merges_zenkaku_and_hankaku_labels():
    """半角 CN1 と全角 ＣＮ１ は同じラベルとして1行に合算される（ユーザー指定、
    2026-07-03: 出力ファイルのラベルはすべて半角にして集計・記録する）。"""
    results = {
        '/tmp/z.dxf': (['CN1', 'ＣＮ１', 'ＣＮ１'], {
            'filename': 'z.dxf',
            'final_count': 3,
            'main_drawing_number': '',
            'source_drawing_number': '',
            'title': '',
            'subtitle': '',
        }),
    }
    wb = _load_wb(create_excel_output(results, 'asc'))
    ws = wb['Total']
    rows = {ws.cell(row=i, column=1).value: ws.cell(row=i, column=2).value
            for i in range(2, ws.max_row + 1)}
    assert rows == {'CN1': 3}


def test_build_region_results_normalizes_names_and_labels():
    """build_region_results は領域名・ラベルとも半角へ正規化して集計する
    （合成データによる単体テスト。半角と全角の同一ラベルが1行に合算される）。"""
    from utils.region_detector import build_region_results
    square = [(0.0, 0.0), (100.0, 0.0), (100.0, 100.0), (0.0, 100.0)]
    analyses = {
        'z.dxf': {
            'frames': [(0, 100, 0, 100)],
            'labels': [('ＣＮ１', 10, 10), ('CN1', 20, 20), ('ＧＮＤ', 30, 30)],
            'regions': [{
                'id': 0, 'frame': 0, 'polygon': square, 'area_pct': 50.0,
                'name_candidates': [(2.5, 'ＳＹＳＴＥＭ　Ｉ／Ｆ　ＢＯＸ')],
            }],
        }
    }
    selections = {('z.dxf', 0): ['ＳＹＳＴＥＭ　Ｉ／Ｆ　ＢＯＸ']}
    rr = build_region_results(analyses, selections, 'asc')
    data = rr['z.dxf']
    assert [r['name'] for r in data['named']] == ['SYSTEM I/F BOX']
    rows = {r['ラベル']: (r['個数'], r['領域']) for r in data['rows']}
    assert rows == {
        'CN1': (2, 'SYSTEM I/F BOX'),   # 全角・半角が合算される
        'GND': (1, 'SYSTEM I/F BOX'),
    }


def test_zenkaku_circuit_symbol_recognized_by_filter():
    """全角の機器符号（ＣＮ１）も filter_non_circuit_symbols が機器符号として
    認識する（判定は半角相当・返り値は元の表記のまま）。"""
    from utils.common_utils import filter_non_circuit_symbols
    matched, excluded = filter_non_circuit_symbols(['ＣＮ１', 'これは説明文'])
    assert matched == ['ＣＮ１']
    assert excluded == 1


# ---------------------------------------------------------------------------
# 「領域一覧」列名変更（「図面」→「ページ No.」）・「領域別ラベル一覧」新設
# ---------------------------------------------------------------------------

def test_region_list_sheet_uses_page_no_header():
    """「領域一覧」シートの列見出しが「ページ No.」であること（旧「図面」から変更）。
    中身はファイル内の図面枠の通し番号のままで、DXF図番ではない。"""
    wb = _load_wb(create_region_excel_output(_make_region_results()))
    ws = wb['領域一覧']
    headers = [c.value for c in ws[1]]
    assert 'ページ No.' in headers
    assert '図面' not in headers
    col = headers.index('ページ No.') + 1
    assert ws.cell(row=2, column=col).value == 1   # frame(=0) + 1


def _make_multi_file_region_results():
    """2ファイルが同じ領域名 'RACK1' を共有し、ラベルが一部重なる合成データ。
    RACK1: R001 は両ファイルに登場（合計 8 = 5+3）、GND は file1 のみ（file2 は 0）。
    file2 には drawing_number が無い（未抽出）ためファイル名にフォールバックする。"""
    return {
        'a.dxf': {
            'rows': [], 'named': [
                {'polygon': [], 'name': 'RACK1', 'id': 0,
                 'frame': 0, 'area_pct': 30.0, 'label_count': 6},
            ],
            'frames': 1, 'regions_detected': 1, 'regions_named': 1,
            'total_in_frame': 6, 'filtered_count': 0, 'final_count': 6,
            'in_region_count': 6,
            'drawing_number': 'EE6492-039-38A',
            'region_label_counts': {'RACK1': {'R001': 5, 'GND': 1}},
        },
        'b.dxf': {
            'rows': [], 'named': [
                {'polygon': [], 'name': 'RACK1', 'id': 0,
                 'frame': 0, 'area_pct': 30.0, 'label_count': 3},
            ],
            'frames': 1, 'regions_detected': 1, 'regions_named': 1,
            'total_in_frame': 3, 'filtered_count': 0, 'final_count': 3,
            'in_region_count': 3,
            'drawing_number': '',   # 未抽出 → ファイル名にフォールバック
            'region_label_counts': {'RACK1': {'R001': 3}},
        },
    }


def test_build_region_label_summary_aggregates_across_files():
    from utils.region_detector import build_region_label_summary
    results = _make_multi_file_region_results()
    files, rows = build_region_label_summary(results)

    assert files == [('a.dxf', 'EE6492-039-38A'), ('b.dxf', 'b')]

    by_label = {r['ラベル']: r for r in rows if r['領域名'] == 'RACK1'}
    assert by_label['R001']['合計個数'] == 8
    assert by_label['R001']['per_file'] == {'a.dxf': 5, 'b.dxf': 3}
    assert by_label['GND']['合計個数'] == 1
    assert by_label['GND']['per_file'].get('b.dxf', 0) == 0   # file2 に無いラベルは0扱い


def test_create_region_excel_output_label_summary_sheet():
    """「領域別ラベル一覧」シートが「領域一覧」の直後に作成され、指定のヘッダー・
    データ配置になっていること。ファイルに存在しないラベルの個数は 0 で埋まる。"""
    results = _make_multi_file_region_results()
    wb = _load_wb(create_region_excel_output(results))

    assert wb.sheetnames.index('領域別ラベル一覧') == wb.sheetnames.index('領域一覧') + 1

    ws = wb['領域別ラベル一覧']
    header = [c.value for c in ws[1]]
    assert header == ['領域名', 'ラベル', '合計個数', '図番', '個数', '図番', '個数']

    rows = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows[row[1]] = row  # ラベルをキーに

    r001 = rows['R001']
    assert r001[0] == 'RACK1'
    assert r001[2] == 8
    assert r001[3] == 'EE6492-039-38A' and r001[4] == 5
    assert r001[5] == 'b' and r001[6] == 3

    gnd = rows['GND']
    assert gnd[2] == 1
    assert gnd[3] == 'EE6492-039-38A' and gnd[4] == 1
    assert gnd[5] == 'b' and gnd[6] == 0   # file2 に無いラベルは 0
