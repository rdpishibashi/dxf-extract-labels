"""tools/reference_designator_analyzer.py の回帰テスト。

パターン・除外リストの定義自体は utils/ref_designator.py（DXF-extract-labels
本体アプリの判定ロジック）を単一の正として参照するため、ここでは集計・
パターン表記生成・Excel 出力の配線を検証する（定義の重複検証はしない）。
"""
import os
import sys
from io import BytesIO

import openpyxl
import pytest

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, PROJECT_ROOT)

from tools import reference_designator_analyzer as rda  # noqa: E402


# ---------------------------------------------------------------------------
# build_pattern_signature
# ---------------------------------------------------------------------------

def test_pattern_signature_letters_only():
    assert rda.build_pattern_signature('ACTAPEFEM', 'letters_only') == 'ACTAPEFEM'


def test_pattern_signature_no_tail_collapses_to_generic_digits():
    assert rda.build_pattern_signature('ACTAPEFEM2', 'letters_digits_any') == 'ACTAPEFEM+1*'


def test_pattern_signature_with_tail_keeps_digit_count():
    assert rda.build_pattern_signature('AAC1B4-07', 'hyphen_letters_digits_any') is None
    # AAC1B4-07 はハイフンパターンではなく letters_digits_any（先頭がAAC+1桁の数字）
    assert rda.build_pattern_signature('AAC1B4-07', 'letters_digits_any') == 'AAC1+A*1*-1*'


def test_pattern_signature_hyphen_pattern():
    assert rda.build_pattern_signature('CN-IF2-1', 'hyphen_letters_digits_any') == 'CN-IF1+-1*'


# ---------------------------------------------------------------------------
# aggregate_labels / classify_aggregated_labels（合成 xlsx）
# ---------------------------------------------------------------------------

def _make_total_xlsx(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Total'
    ws.append(['ラベル', '個数'])
    for label, count in rows:
        ws.append([label, count])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_aggregate_labels_merges_zenkaku_hankaku_and_sums_across_files():
    src_a = _make_total_xlsx([('R10', 3), ('ＲＡＣＫ１', 2)])
    src_b = _make_total_xlsx([('R10', 4), ('GND', 1)])
    agg, stats = rda.aggregate_labels([('a.xlsx', src_a), ('b.xlsx', src_b)])
    assert agg['R10']['count'] == 7
    assert agg['R10']['files'] == {'a.xlsx', 'b.xlsx'}
    assert agg['RACK1']['count'] == 2   # 全角→半角正規化
    assert stats == [
        {'ファイル名': 'a.xlsx', 'ラベル種類数': 2},
        {'ファイル名': 'b.xlsx', 'ラベル種類数': 2},
    ]


def test_classify_aggregated_labels_splits_candidate_excluded_confirmed_and_drops_no_match():
    src = _make_total_xlsx([
        ('R10', 5), ('D100', 4), ('C1A', 6), ('KZ3', 2),
        ('GND', 9), ('(2/5)', 1), ('AWG14', 3),
    ])
    agg, _ = rda.aggregate_labels([('x.xlsx', src)])
    ref_rows, sig_rows, exc_impact, conf_impact = rda.classify_aggregated_labels(agg)

    by_label = {r['ラベル']: r for r in ref_rows}
    assert '(2/5)' not in by_label   # 3パターンいずれにも不一致 -> 完全に除外
    assert by_label['GND']['除外ステータス'] == '確定'
    assert by_label['GND']['除外カテゴリ'] == 'circuit_description'
    assert by_label['AWG14']['除外カテゴリ'] == 'wire_gauge'

    # R10 / D100: 英字+数字2桁 or 3桁 -> 確定パターン
    assert by_label['R10']['確定ステータス'] == '確定'
    assert by_label['R10']['確定カテゴリ'] == 'letters_digits_2or3'
    assert by_label['D100']['確定カテゴリ'] == 'letters_digits_2or3'
    # C1A: 数字が1桁のみなので letters_digits_2or3_letter（2-3桁要求）には一致しない
    assert by_label['C1A']['確定ステータス'] == ''
    # KZ3: 英字2文字+数字1桁 -> 確定パターンいずれにも一致せず未確定のまま
    # （CNプレフィックスの cn_single_digit 等は対象外のプレフィックスのため無関係）
    assert by_label['KZ3']['確定ステータス'] == ''
    assert by_label['KZ3']['除外ステータス'] == ''

    assert exc_impact['circuit_description'] == {'labels': 1, 'count': 9}
    assert exc_impact['wire_gauge'] == {'labels': 1, 'count': 3}
    assert conf_impact['letters_digits_2or3'] == {'labels': 2, 'count': 9}   # R10(5)+D100(4)


# ---------------------------------------------------------------------------
# matched_confirmed_category
# ---------------------------------------------------------------------------

@pytest.mark.parametrize('judgment,expected', [
    ('R10', 'letters_digits_2or3'),
    ('R100', 'letters_digits_2or3'),
    ('CN3', 'cn_single_digit'),   # CN+数字1桁（2026-07-10 追加、他パターンより先に判定）
    ('R10A', 'letters_digits_2or3_letter'),
    ('CN-IF21', 'cn_if_prefix'),   # "CN-IF"+任意の文字が先に判定される
    ('CN-IF2-1', 'cn_if_prefix'),  # 2026-07-10: 末尾に続きがあっても確定対象
    ('D1', 'single_letter_digits_except_ab'),   # 単一英字+数字は桁数を問わない
    ('D1000', 'single_letter_digits_except_ab'),
    # A,B は「単一英字+数字」パターン(4)の対象外だが、複数桁ルール(1,2)には
    # A,B除外は無い（実運用では A1/B12 等は既存の terminal_row_letter_digit
    # 除外パターンで先に除外されるため、確定パターン判定まで到達しない）。
    ('D123', 'letters_digits_2or3'),   # パターン1が先勝ち（3桁は letters_digits_2or3 にも一致）
    ('B12', 'letters_digits_2or3'),
])
def test_matched_confirmed_category(judgment, expected):
    assert rda.matched_confirmed_category(judgment) == expected


def test_single_letter_except_ab_matches_any_digit_length():
    """パターン4（A,B以外の1英大文字+数字）は桁数を問わない
    （パターン1,2の「2桁または3桁」制限はパターン4には適用されない）。"""
    assert rda.matched_confirmed_category('R1') == 'single_letter_digits_except_ab'
    assert rda.matched_confirmed_category('R1000') == 'single_letter_digits_except_ab'


def test_a_and_b_excluded_only_from_single_letter_pattern():
    """A,B の除外はパターン4（単一英字+数字）専用。他の確定パターンには適用されない
    （A1/B12等はterminal_row_letter_digit除外で候補プールに残らないため実害はない）。"""
    assert rda.matched_confirmed_category('A1') is None
    assert rda.matched_confirmed_category('B1') is None


# ---------------------------------------------------------------------------
# build_output_workbook
# ---------------------------------------------------------------------------

def test_build_output_workbook_sheet_names_and_remaining_sorted_desc():
    src = _make_total_xlsx([('CNX', 2), ('CNY', 9), ('GND', 5)])
    agg, _ = rda.aggregate_labels([('x.xlsx', src)])
    ref_rows, sig_rows, exc_impact, conf_impact = rda.classify_aggregated_labels(agg)
    data = rda.build_output_workbook(ref_rows, sig_rows, exc_impact, conf_impact)

    wb = openpyxl.load_workbook(BytesIO(data))
    assert wb.sheetnames == [
        'ReferenceDesignators', 'Patterns', 'PatternSignatures',
        'ExclusionPatterns', 'ConfirmedPatterns', 'ConfirmedDesignators',
        'RemainingUnclassified',
    ]

    remaining = list(wb['RemainingUnclassified'].iter_rows(min_row=2, values_only=True))
    assert [r[0] for r in remaining] == ['CNY', 'CNX']   # 個数降順、GNDは含まれない

    ref_labels = {r[0] for r in wb['ReferenceDesignators'].iter_rows(min_row=2, values_only=True)}
    assert ref_labels == {'CNX', 'CNY', 'GND'}


def test_build_output_workbook_confirmed_labels_excluded_from_remaining():
    """確定パターンに一致したラベル（R10）は RemainingUnclassified から除かれ、
    ConfirmedPatterns シートの該当ラベル数・個数に反映される。"""
    src = _make_total_xlsx([('R10', 4), ('CNX', 2), ('GND', 5)])
    agg, _ = rda.aggregate_labels([('x.xlsx', src)])
    ref_rows, sig_rows, exc_impact, conf_impact = rda.classify_aggregated_labels(agg)
    data = rda.build_output_workbook(ref_rows, sig_rows, exc_impact, conf_impact)
    wb = openpyxl.load_workbook(BytesIO(data))

    remaining_labels = {r[0] for r in wb['RemainingUnclassified'].iter_rows(min_row=2, values_only=True)}
    assert remaining_labels == {'CNX'}

    confirmed_rows = {r[0]: r for r in wb['ConfirmedPatterns'].iter_rows(min_row=2, values_only=True)}
    row = confirmed_rows['letters_digits_2or3']
    # 列構成: カテゴリ,ステータス,種別,判定基準,パターン/一覧,理由,該当ラベル数,該当個数合計
    assert row[6] == 1   # 該当ラベル数
    assert row[7] == 4   # 該当個数合計

    designator_rows = list(wb['ConfirmedDesignators'].iter_rows(min_row=2, values_only=True))
    assert designator_rows == [('R10', 4, 1, 'letters_digits_2or3')]


def test_build_output_workbook_empty_input_does_not_raise():
    data = rda.build_output_workbook([], [], {}, {})
    wb = openpyxl.load_workbook(BytesIO(data))
    assert 'ReferenceDesignators' in wb.sheetnames
    assert 'ConfirmedPatterns' in wb.sheetnames
    assert 'ConfirmedDesignators' in wb.sheetnames


# ---------------------------------------------------------------------------
# _iter_input_sources（フォルダ探索）
# ---------------------------------------------------------------------------

def test_iter_input_sources_combines_uploads_and_folder(tmp_path):
    (tmp_path / 'extracted_labels.xlsx').write_bytes(b'dummy')
    (tmp_path / 'extracted_labels (1).xlsx').write_bytes(b'dummy')
    (tmp_path / '~$extracted_labels.xlsx').write_bytes(b'dummy')   # ロックファイル除外
    (tmp_path / 'other.xlsx').write_bytes(b'dummy')                # パターン非一致除外

    class _FakeUpload:
        def __init__(self, name, content):
            self.name = name
            self._content = content

        def getvalue(self):
            return self._content

    uploads = [_FakeUpload('uploaded.xlsx', b'uploaded-bytes')]
    sources = rda._iter_input_sources(uploads, [str(tmp_path)], 'extracted_labels*.xlsx', False)
    names = sorted(n for n, _ in sources)
    assert names == ['extracted_labels (1).xlsx', 'extracted_labels.xlsx', 'uploaded.xlsx']


def test_iter_input_sources_missing_folder_is_skipped_not_raised():
    sources = rda._iter_input_sources([], ['/no/such/folder'], 'extracted_labels*.xlsx', False)
    assert sources == []


# ---------------------------------------------------------------------------
# 判断ログ集計（aggregate_decision_log / build_decision_log_suggestions /
# build_decision_log_workbook）v1.7.0
# ---------------------------------------------------------------------------

_LOG_HEADER = ','.join([
    'timestamp', 'source', 'file_name', 'drawing_number',
    'label', 'decision', 'count', 'app_version', 'patterns_version',
])


def _make_log_csv(rows):
    """rows: [(file_name, label, decision, count), ...] から CSV バイト列を作る。"""
    lines = [_LOG_HEADER]
    for file_name, label, decision, count in rows:
        lines.append(
            f"2026-07-10T10:00:00,cloud,{file_name},,{label},{decision},{count},1.7.0,1.6.4")
    return ('\n'.join(lines) + '\n').encode('utf-8-sig')


def test_aggregate_decision_log_sums_adopted_and_rejected_across_sources():
    csv_a = _make_log_csv([
        ('a.dxf', 'R10', 'adopted', 2),
        ('a.dxf', 'CN3', 'rejected', 1),
    ])
    csv_b = _make_log_csv([
        ('b.dxf', 'R10', 'adopted', 1),
        ('b.dxf', 'CN3', 'rejected', 3),
    ])
    agg, stats = rda.aggregate_decision_log([('a.csv', csv_a), ('b.csv', csv_b)])

    assert agg['R10']['adopted'] == 3
    assert agg['R10']['rejected'] == 0
    assert agg['R10']['files'] == {'a.dxf', 'b.dxf'}
    assert agg['CN3']['adopted'] == 0
    assert agg['CN3']['rejected'] == 4
    assert stats == [
        {'ソース': 'a.csv', '行数': 2},
        {'ソース': 'b.csv', '行数': 2},
    ]


def test_aggregate_decision_log_normalizes_zenkaku_labels():
    csv_data = _make_log_csv([('a.dxf', 'ＲＡＣＫ１', 'adopted', 1)])
    agg, _ = rda.aggregate_decision_log([('a.csv', csv_data)])
    assert 'RACK1' in agg
    assert agg['RACK1']['adopted'] == 1


def test_aggregate_decision_log_skips_malformed_source_without_raising():
    bad = b'not,a,valid,decision,log\n1,2,3,4,5'
    agg, stats = rda.aggregate_decision_log([('bad.csv', bad)])
    assert agg == {}
    assert stats == [{'ソース': 'bad.csv', '行数': 0}]


def test_aggregate_decision_log_accepts_already_fetched_text():
    text = _make_log_csv([('a.dxf', 'R10', 'adopted', 1)]).decode('utf-8-sig')
    agg, _ = rda.aggregate_decision_log([('GitHub:org/repo', text)])
    assert agg['R10']['adopted'] == 1


def test_build_decision_log_suggestions_classifies_by_rate_and_min_occurrences():
    agg = {
        'ALWAYS_ADOPTED': {'adopted': 5, 'rejected': 0, 'files': {'a.dxf'}},
        'ALWAYS_REJECTED': {'adopted': 0, 'rejected': 5, 'files': {'a.dxf'}},
        'MIXED': {'adopted': 3, 'rejected': 3, 'files': {'a.dxf'}},
        'TOO_FEW': {'adopted': 1, 'rejected': 0, 'files': {'a.dxf'}},
    }
    rows = rda.build_decision_log_suggestions(
        agg, min_occurrences=3, confirm_rate=1.0, exclude_rate=1.0)
    by_label = {r['ラベル']: r for r in rows}

    assert by_label['ALWAYS_ADOPTED']['提案'] == '確定パターン候補'
    assert by_label['ALWAYS_REJECTED']['提案'] == '除外パターン候補'
    assert by_label['MIXED']['提案'] == '様子見'
    assert by_label['TOO_FEW']['提案'] == '様子見（サンプル不足）'

    # 合計降順・同数ならラベル昇順でソートされる（MIXED=6, ALWAYS_*=5, TOO_FEW=1）
    assert [r['ラベル'] for r in rows] == [
        'MIXED', 'ALWAYS_ADOPTED', 'ALWAYS_REJECTED', 'TOO_FEW']


def test_build_decision_log_suggestions_rate_thresholds_are_configurable():
    agg = {'MOSTLY_ADOPTED': {'adopted': 9, 'rejected': 1, 'files': {'a.dxf'}}}
    strict = rda.build_decision_log_suggestions(
        agg, min_occurrences=3, confirm_rate=1.0, exclude_rate=1.0)
    assert strict[0]['提案'] == '様子見'

    lenient = rda.build_decision_log_suggestions(
        agg, min_occurrences=3, confirm_rate=0.8, exclude_rate=1.0)
    assert lenient[0]['提案'] == '確定パターン候補'


def test_build_decision_log_workbook_contains_summary_sheet():
    rows = [{
        'ラベル': 'R10', '採用数': 5, '非採用数': 0, '合計': 5,
        '採用率': 1.0, '出現ファイル数': 1, '提案': '確定パターン候補',
    }]
    data = rda.build_decision_log_workbook(rows)
    wb = openpyxl.load_workbook(BytesIO(data))
    assert wb.sheetnames == ['DecisionLogSummary']
    values = list(wb['DecisionLogSummary'].iter_rows(min_row=2, values_only=True))
    assert values == [('R10', 5, 0, 5, 1.0, 1, '確定パターン候補')]


def test_build_decision_log_workbook_empty_input_does_not_raise():
    data = rda.build_decision_log_workbook([])
    wb = openpyxl.load_workbook(BytesIO(data))
    assert 'DecisionLogSummary' in wb.sheetnames
