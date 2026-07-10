"""Reference Designator 抽出検討ツール

extracted_labels*.xlsx（DXF-extract-labels の「機器符号（候補）以外も抽出」ON
で出力した Excel。`Total` シートに ラベル・個数 列を持つ）を複数入力し、
reference_designator_candidates.xlsx と同じ構成の分析用 Excel を生成する。

パターン・除外リストの定義は `utils/ref_designator.py` を単一の正として再利用
する（本体アプリの判定ロジックと乖離しないため、ここで独自に定義し直さない）。

起動方法:
    streamlit run tools/reference_designator_analyzer.py
"""
import csv
import glob
import os
import re
import sys
from collections import Counter, defaultdict
from io import BytesIO, StringIO

import openpyxl
import streamlit as st

_CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.dirname(_CURRENT_DIR)
sys.path.insert(0, _PROJECT_ROOT)

from utils import ref_designator as rd  # noqa: E402
from utils import decision_log as dlog  # noqa: E402

st.set_page_config(
    page_title="Reference Designator 抽出検討ツール",
    page_icon="🔍",
    layout="wide",
)


# ============================================================
# 入力ファイルの収集・読み込み
# ============================================================

def _iter_input_sources(uploaded_files, folder_paths, glob_pattern, recursive):
    """(表示名, バイト列 or ローカルパス) のリストを返す。"""
    sources = []
    for uf in uploaded_files or []:
        sources.append((uf.name, uf.getvalue()))

    for folder in folder_paths:
        folder = folder.strip()
        if not folder:
            continue
        if not os.path.isdir(folder):
            st.warning(f"フォルダが見つかりません: {folder}")
            continue
        pattern_path = os.path.join(folder, '**', glob_pattern) if recursive \
            else os.path.join(folder, glob_pattern)
        found = glob.glob(pattern_path, recursive=recursive)
        for path in sorted(found):
            base = os.path.basename(path)
            if base.startswith('~$'):
                continue
            sources.append((base, path))
    return sources


def _read_total_sheet(name, source):
    """(表示名, バイト列またはパス) から Total シートの (ラベル, 個数) を読む。"""
    try:
        target = BytesIO(source) if isinstance(source, (bytes, bytearray)) else source
        wb = openpyxl.load_workbook(target, read_only=True, data_only=True)
    except Exception as e:
        st.warning(f"{name}: 読み込みに失敗しました（{e}）")
        return []
    if 'Total' not in wb.sheetnames:
        st.warning(f"{name}: Total シートが見つかりません（スキップ）")
        wb.close()
        return []
    ws = wb['Total']
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        label = str(row[0])
        count = int(row[1]) if row[1] is not None else 0
        rows.append((label, count))
    wb.close()
    return rows


def aggregate_labels(sources):
    """各ソースの Total シートを正規化して集計する。

    戻り値: (agg, per_source_stats)
      agg: {正規化ラベル: {'count': int, 'files': set(表示名)}}
      per_source_stats: [{'ファイル名': str, 'ラベル種類数': int}]
    """
    agg = defaultdict(lambda: {'count': 0, 'files': set()})
    per_source_stats = []
    for name, source in sources:
        rows = _read_total_sheet(name, source)
        per_file = Counter()
        for label, count in rows:
            norm = rd.normalize_label(label)
            if not norm:
                continue
            per_file[norm] += count
        for norm, count in per_file.items():
            agg[norm]['count'] += count
            agg[norm]['files'].add(name)
        per_source_stats.append({'ファイル名': name, 'ラベル種類数': len(per_file)})
    return agg, per_source_stats


# ============================================================
# パターン表記（電気技術者向け人間可読ノーテーション）
# ============================================================
#
# 「数字列の前の部分はそのまま」「直後の数字は桁数表記(1,12,123...)」「その後ろは
# +英字繰返し=A*/数字繰返し=1*/ハイフンはそのまま、後ろに何も続かなければ+1*に
# 集約」という表記ルール（2026-07 ユーザーと確定）。本体アプリでは使わない
# 分析専用の表記のため、ref_designator.py には持ち込まずここに閉じる。

_ASCENDING_DIGITS = '123456789'
_TAIL_TOKEN_RE = re.compile(r'[A-Z]+|[0-9]+|-')
_PREFIX_DIGITS_TAIL_HYPHEN_RE = re.compile(r'^([A-Z]+-[A-Z]+)([0-9]+)([A-Z0-9-]*)$')
_PREFIX_DIGITS_TAIL_RE = re.compile(r'^([A-Z]+)([0-9]+)([A-Z0-9-]*)$')


def _ascending_digit_repr(n: int) -> str:
    if n <= len(_ASCENDING_DIGITS):
        return _ASCENDING_DIGITS[:n]
    return ''.join(str((i % 9) + 1) for i in range(n))


def _encode_tail(tail: str) -> str:
    out = []
    for tok in _TAIL_TOKEN_RE.findall(tail):
        if tok == '-':
            out.append('-')
        elif tok.isdigit():
            out.append('1*')
        else:
            out.append('A*')
    return ''.join(out)


def build_pattern_signature(label: str, pattern_name: str):
    """ラベルとその一致パターン名から人間可読のパターン表記を作る。"""
    if pattern_name == 'letters_only':
        return label
    rx = _PREFIX_DIGITS_TAIL_HYPHEN_RE if pattern_name == 'hyphen_letters_digits_any' \
        else _PREFIX_DIGITS_TAIL_RE
    m = rx.match(label)
    if not m:
        return None
    prefix, digits, tail = m.groups()
    if tail == '':
        return f'{prefix}+1*'
    return f'{prefix}{_ascending_digit_repr(len(digits))}+{_encode_tail(tail)}'


# ============================================================
# 確定パターン（RemainingUnclassified のうち Reference Designator と確定できるもの）
# ============================================================
#
# ExclusionPatterns で明らかに Reference Designator ではないものを除いた後の
# 残り（RemainingUnclassified）から、さらに「確実に Reference Designator と
# 判定してよい」形をユーザーと確定した4パターン（2026-07-10）。本体アプリ
# （`utils/ref_designator.py`）も v1.6.3 でこの4パターンを取り込み、確定した
# ラベルは「未確定ラベル」UI でのレビューを経ずに自動採用するようになったため、
# ここでも `utils/ref_designator.py` を単一の正として参照する（独自に定義し直さない）。

CONFIRMED_PATTERN_CATEGORIES = rd.CONFIRMED_PATTERN_CATEGORIES
matched_confirmed_category = rd.matched_confirmed_category


# ============================================================
# 分類・Excel 生成
# ============================================================

def classify_aggregated_labels(agg):
    """agg（aggregate_labels の戻り値）を分類する。

    戻り値: (ref_designator_rows, signature_rows, exclusion_impact, confirmed_impact)
      ref_designator_rows: [{'ラベル','個数','出現ファイル数','一致パターン',
                              '除外カテゴリ','除外ステータス',
                              '確定カテゴリ','確定ステータス'}]（Patterns一致のみ）
      signature_rows: [{'パターン表記','個数','出現ファイル数','該当ラベル数'}]
      exclusion_impact: {カテゴリ名: {'labels': int, 'count': int}}
      confirmed_impact: {カテゴリ名: {'labels': int, 'count': int}}
        （除外されなかったラベルのうち、確定パターンに一致したものだけを集計）
    """
    ref_rows = []
    sig_agg = defaultdict(lambda: {'count': 0, 'files': set(), 'labels': set()})
    exclusion_impact = defaultdict(lambda: {'labels': 0, 'count': 0})
    confirmed_impact = defaultdict(lambda: {'labels': 0, 'count': 0})

    for label, data in agg.items():
        judgment = rd._judgment_text(label)
        status, category = rd.classify_judgment_detailed(judgment)
        if status == 'no_match':
            continue
        pattern_name = rd.matched_pattern_name(judgment)

        confirmed_category = None
        if status == 'candidate':
            confirmed_category = matched_confirmed_category(label)

        ref_rows.append({
            'ラベル': label,
            '個数': data['count'],
            '出現ファイル数': len(data['files']),
            '一致パターン': pattern_name,
            '除外カテゴリ': category or '',
            '除外ステータス': '確定' if status == 'excluded' else '',
            '確定カテゴリ': confirmed_category or '',
            '確定ステータス': '確定' if confirmed_category else '',
        })
        if status == 'excluded':
            exclusion_impact[category]['labels'] += 1
            exclusion_impact[category]['count'] += data['count']
        elif confirmed_category:
            confirmed_impact[confirmed_category]['labels'] += 1
            confirmed_impact[confirmed_category]['count'] += data['count']

        sig = build_pattern_signature(judgment, pattern_name)
        if sig:
            sig_agg[sig]['count'] += data['count']
            sig_agg[sig]['files'] |= data['files']
            sig_agg[sig]['labels'].add(label)

    ref_rows.sort(key=lambda r: r['ラベル'])
    signature_rows = [
        {'パターン表記': sig, '個数': d['count'], '出現ファイル数': len(d['files']),
         '該当ラベル数': len(d['labels'])}
        for sig, d in sig_agg.items()
    ]
    signature_rows.sort(key=lambda r: r['パターン表記'])
    return ref_rows, signature_rows, exclusion_impact, confirmed_impact


def build_output_workbook(ref_rows, signature_rows, exclusion_impact, confirmed_impact):
    """reference_designator_candidates.xlsx と同じ構成の Excel を作る。"""
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = 'ReferenceDesignators'
    ws1.append(['ラベル', '個数', '出現ファイル数', '一致パターン', '除外カテゴリ', '除外ステータス',
                '確定カテゴリ', '確定ステータス'])
    for r in ref_rows:
        ws1.append([r['ラベル'], r['個数'], r['出現ファイル数'], r['一致パターン'],
                    r['除外カテゴリ'], r['除外ステータス'],
                    r['確定カテゴリ'], r['確定ステータス']])
    for col, width in zip('ABCDEFGH', (20, 10, 14, 26, 26, 12, 28, 12)):
        ws1.column_dimensions[col].width = width
    ws1.freeze_panes = 'A2'

    ws2 = wb.create_sheet('Patterns')
    ws2.append(['パターン名', '正規表現', '説明'])
    for name, rx, desc in rd.PATTERN_CATEGORIES:
        ws2.append([name, rx.pattern, desc])
    ws2.append(['combined（統合）', rd.CANDIDATE_PATTERN.pattern,
                '上記3パターンのいずれかに一致（判定前に NFKC 正規化・前後空白除去・括弧以降除去を適用）'])
    for col, width in zip('ABC', (24, 60, 60)):
        ws2.column_dimensions[col].width = width

    ws3 = wb.create_sheet('PatternSignatures')
    ws3.append(['パターン表記', '個数', '出現ファイル数', '該当ラベル数'])
    for r in signature_rows:
        ws3.append([r['パターン表記'], r['個数'], r['出現ファイル数'], r['該当ラベル数']])
    for col, width in zip('ABCD', (28, 10, 14, 12)):
        ws3.column_dimensions[col].width = width
    ws3.freeze_panes = 'A2'

    ws4 = wb.create_sheet('ExclusionPatterns')
    ws4.append(['カテゴリ', 'ステータス', '種別', 'パターン/一覧', '理由', '該当ラベル数', '該当個数合計'])
    for name, (words, desc) in rd.EXCLUSION_EXACT_CATEGORIES.items():
        imp = exclusion_impact.get(name, {'labels': 0, 'count': 0})
        ws4.append([name, '確定', '完全一致（列挙）', ', '.join(sorted(words)), desc,
                    imp['labels'], imp['count']])
    for name, rx, desc in rd.EXCLUSION_REGEX_CATEGORIES:
        imp = exclusion_impact.get(name, {'labels': 0, 'count': 0})
        ws4.append([name, '確定', '正規表現', rx.pattern, desc, imp['labels'], imp['count']])
    for col, width in zip('ABCDEFG', (26, 10, 22, 60, 45, 12, 14)):
        ws4.column_dimensions[col].width = width
    ws4.freeze_panes = 'A2'

    ws5 = wb.create_sheet('ConfirmedPatterns')
    ws5.append(['カテゴリ', 'ステータス', '種別', '判定基準', 'パターン/一覧', '理由', '該当ラベル数', '該当個数合計'])
    for name, basis, rx, desc in CONFIRMED_PATTERN_CATEGORIES:
        imp = confirmed_impact.get(name, {'labels': 0, 'count': 0})
        ws5.append([name, '確定', '正規表現', basis, rx.pattern, desc, imp['labels'], imp['count']])
    for col, width in zip('ABCDEFGH', (30, 10, 12, 12, 55, 55, 12, 14)):
        ws5.column_dimensions[col].width = width
    ws5.freeze_panes = 'A2'

    ws5b = wb.create_sheet('ConfirmedDesignators')
    ws5b.append(['ラベル', '個数', '出現ファイル数', '一致パターン'])
    confirmed_rows = [r for r in ref_rows if r['確定ステータス'] == '確定']
    confirmed_rows.sort(key=lambda r: -r['個数'])
    for r in confirmed_rows:
        ws5b.append([r['ラベル'], r['個数'], r['出現ファイル数'], r['確定カテゴリ']])
    for col, width in zip('ABCD', (20, 10, 14, 30)):
        ws5b.column_dimensions[col].width = width
    ws5b.freeze_panes = 'A2'

    ws6 = wb.create_sheet('RemainingUnclassified')
    ws6.append(['ラベル', '個数', '出現ファイル数', '一致パターン'])
    remaining = [r for r in ref_rows if r['除外ステータス'] == '' and r['確定ステータス'] == '']
    remaining.sort(key=lambda r: -r['個数'])
    for r in remaining:
        ws6.append([r['ラベル'], r['個数'], r['出現ファイル数'], r['一致パターン']])
    for col, width in zip('ABCD', (20, 10, 14, 20)):
        ws6.column_dimensions[col].width = width
    ws6.freeze_panes = 'A2'

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ============================================================
# 判断ログ（decision_log.csv）の集計・パターン候補提案（v1.7.0）
# ============================================================
#
# DXF-extract-labels 本体アプリの「未確定ラベル」UI でユーザーが行った採用/非採用の
# 判断（utils/decision_log.py が記録）を集計し、確定パターン・除外パターンの
# 候補を機械的に提案する。GitHub 上のログ専用リポジトリから直接取得するか、
# ローカル/Dropbox の decision_log.csv をアップロード・フォルダ指定で読み込む。

def _read_decision_log_rows(name, source):
    """(表示名, バイト列・パス・テキストのいずれか) から判断ログの行（dict）を読む。"""
    try:
        if isinstance(source, (bytes, bytearray)):
            text = source.decode('utf-8-sig')
        elif isinstance(source, str) and os.path.exists(source):
            with open(source, encoding='utf-8-sig') as f:
                text = f.read()
        else:
            text = source  # GitHub から取得済みのテキストそのもの
    except Exception as e:
        st.warning(f"{name}: 読み込みに失敗しました（{e}）")
        return []
    if not text:
        return []
    reader = csv.DictReader(StringIO(text))
    if not reader.fieldnames or not {'label', 'decision'}.issubset(reader.fieldnames):
        st.warning(f"{name}: 判断ログの形式ではありません（label/decision 列が見つかりません）")
        return []
    return list(reader)


def aggregate_decision_log(sources):
    """判断ログ（複数ソース）を正規化ラベル単位で集計する。

    戻り値: (agg, per_source_stats)
      agg: {正規化ラベル: {'adopted': int, 'rejected': int, 'files': set(file_name)}}
      per_source_stats: [{'ソース': str, '行数': int}]
    """
    agg = defaultdict(lambda: {'adopted': 0, 'rejected': 0, 'files': set()})
    per_source_stats = []
    for name, source in sources:
        rows = _read_decision_log_rows(name, source)
        for row in rows:
            label = rd.normalize_label(row.get('label', ''))
            decision = row.get('decision', '')
            if not label or decision not in ('adopted', 'rejected'):
                continue
            try:
                count = int(row.get('count') or 0)
            except ValueError:
                count = 0
            agg[label][decision] += count
            fname = row.get('file_name') or ''
            if fname:
                agg[label]['files'].add(fname)
        per_source_stats.append({'ソース': name, '行数': len(rows)})
    return agg, per_source_stats


def build_decision_log_suggestions(agg, min_occurrences=3, confirm_rate=1.0, exclude_rate=1.0):
    """集計結果から確定/除外パターンの候補を提案する。

    - 合計出現回数（adopted+rejected）が min_occurrences 未満の行はサンプル不足として
      「様子見」に固定する（提案対象外）。
    - 採用率（adopted/合計） >= confirm_rate → 「確定パターン候補」
    - 非採用率（rejected/合計） >= exclude_rate → 「除外パターン候補」
    - どちらでもなければ「様子見」
    """
    rows = []
    for label, d in agg.items():
        total = d['adopted'] + d['rejected']
        if total == 0:
            continue
        adopt_rate = d['adopted'] / total
        reject_rate = d['rejected'] / total
        if total < min_occurrences:
            suggestion = '様子見（サンプル不足）'
        elif adopt_rate >= confirm_rate:
            suggestion = '確定パターン候補'
        elif reject_rate >= exclude_rate:
            suggestion = '除外パターン候補'
        else:
            suggestion = '様子見'
        rows.append({
            'ラベル': label,
            '採用数': d['adopted'],
            '非採用数': d['rejected'],
            '合計': total,
            '採用率': round(adopt_rate, 3),
            '出現ファイル数': len(d['files']),
            '提案': suggestion,
        })
    rows.sort(key=lambda r: (-r['合計'], r['ラベル']))
    return rows


def build_decision_log_workbook(rows):
    """判断ログ集計結果を1シートの Excel にする。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'DecisionLogSummary'
    ws.append(['ラベル', '採用数', '非採用数', '合計', '採用率', '出現ファイル数', '提案'])
    for r in rows:
        ws.append([r['ラベル'], r['採用数'], r['非採用数'], r['合計'],
                    r['採用率'], r['出現ファイル数'], r['提案']])
    for col, width in zip('ABCDEFG', (20, 10, 10, 10, 10, 14, 22)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = 'A2'
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ============================================================
# Streamlit UI
# ============================================================

def _app_extracted_labels():
    st.write(
        "複数の `extracted_labels*.xlsx`（Total シート）からラベルを集計し、"
        "機器符号（候補）パターン・除外パターンの判定結果を "
        "`reference_designator_candidates.xlsx` と同じ構成の Excel にまとめます。"
    )
    with st.expander("ℹ️ このツールについて", expanded=False):
        st.info(
            "DXF-extract-labels 本体の「機器符号（候補）以外も抽出」ON で出力した "
            "extracted_labels*.xlsx を入力とし、Reference Designator の抽出パターン・"
            "除外パターンを検討するための分析用ツールです。\n\n"
            "パターン・除外リストの定義は `utils/ref_designator.py`（本体アプリの判定"
            "ロジック）を単一の正として参照するため、本体の挙動とここでの分析結果は"
            "常に一致します。\n\n"
            "**本体アプリとの違い**: 本体アプリは DXF ファイルを直接解析し、図面枠・"
            "図面情報欄（フォーマットブロック）を構造的に判定して除外します。本ツールは"
            "集計済みラベル一覧（Total シート、座標情報なし）だけを見るため、その構造的"
            "除外は行えません。人名等は個別リストを持たない設計のため、本ツールの結果には"
            "図面情報欄由来のラベル（人名等）がパターン一致・除外非該当のまま残ることが"
            "あります（本体アプリの実際の出力ではこれらは構造的除外で消えます）。"
        )

    st.subheader("入力")
    col_left, col_right = st.columns(2)
    with col_left:
        uploaded_files = st.file_uploader(
            "extracted_labels*.xlsx をアップロード（複数可）",
            type="xlsx",
            accept_multiple_files=True,
        )
    with col_right:
        folder_text = st.text_area(
            "フォルダパス（1行に1つ、ローカルのフルパス）",
            value="",
            help="このマシン上のディレクトリを指定すると、配下の xlsx ファイルも"
                 "対象に加えます。アップロードと併用できます。",
        )
        glob_pattern = st.text_input("検索パターン（glob）", value="extracted_labels*.xlsx")
        recursive = st.checkbox("サブフォルダも検索する", value=True)

    folder_paths = [line for line in folder_text.splitlines() if line.strip()]

    if st.button("候補を抽出", type="primary"):
        sources = _iter_input_sources(uploaded_files, folder_paths, glob_pattern, recursive)
        if not sources:
            st.warning("入力ファイルがありません。アップロードするか、フォルダを指定してください。")
        else:
            with st.spinner(f"{len(sources)} 個のファイルを処理中..."):
                agg, per_source_stats = aggregate_labels(sources)
                ref_rows, signature_rows, exclusion_impact, confirmed_impact = \
                    classify_aggregated_labels(agg)
                excel_bytes = build_output_workbook(
                    ref_rows, signature_rows, exclusion_impact, confirmed_impact)

            st.session_state['rda_result'] = {
                'excel_bytes': excel_bytes,
                'sources': [name for name, _ in sources],
                'per_source_stats': per_source_stats,
                'total_labels': len(agg),
                'pattern_matched': len(ref_rows),
                'excluded': sum(1 for r in ref_rows if r['除外ステータス']),
                'confirmed': sum(1 for r in ref_rows if r['確定ステータス']),
                'remaining': sum(1 for r in ref_rows
                                  if not r['除外ステータス'] and not r['確定ステータス']),
                'exclusion_impact': dict(exclusion_impact),
                'confirmed_impact': dict(confirmed_impact),
                'signature_count': len(signature_rows),
            }
            st.rerun()

    result = st.session_state.get('rda_result')
    if result:
        st.subheader("結果")
        st.success(f"{len(result['sources'])} 個のファイルを処理しました。")

        m1, m2, m3, m4, m5 = st.columns(5)
        m1.metric("正規化後ラベル種類数", result['total_labels'])
        m2.metric("Patterns 一致", result['pattern_matched'])
        m3.metric("除外", result['excluded'])
        m4.metric("確定（機器符号）", result['confirmed'])
        m5.metric("未分類（要検討）", result['remaining'])

        with st.expander("📊 入力ファイル別の内訳", expanded=False):
            import pandas as pd
            st.dataframe(pd.DataFrame(result['per_source_stats']), width='stretch', hide_index=True)

        with st.expander("📊 除外カテゴリ別の内訳", expanded=False):
            import pandas as pd
            rows = [
                {'カテゴリ': name, '該当ラベル数': imp['labels'], '該当個数合計': imp['count']}
                for name, imp in sorted(result['exclusion_impact'].items(),
                                         key=lambda kv: -kv[1]['labels'])
            ]
            st.dataframe(pd.DataFrame(rows), width='stretch', hide_index=True)

        with st.expander("📊 確定パターン別の内訳", expanded=False):
            import pandas as pd
            rows = [
                {'カテゴリ': name, '該当ラベル数': imp['labels'], '該当個数合計': imp['count']}
                for name, imp in sorted(result['confirmed_impact'].items(),
                                         key=lambda kv: -kv[1]['labels'])
            ]
            st.dataframe(pd.DataFrame(rows), width='stretch', hide_index=True)

        st.download_button(
            label="Excelをダウンロード（reference_designator_candidates.xlsx 形式）",
            data=result['excel_bytes'],
            file_name="reference_designator_candidates.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
        )

        if st.button("🔄 リセット"):
            del st.session_state['rda_result']
            st.rerun()


def _app_decision_log():
    st.write(
        "DXF-extract-labels 本体アプリの「未確定ラベル」UI でユーザーが行った"
        "採用/非採用の判断ログ（`decision_log.csv`）を集計し、確定パターン・"
        "除外パターンの候補を提案します。"
    )
    with st.expander("ℹ️ このモードについて", expanded=False):
        st.info(
            "判断ログは本体アプリが GitHub のログ専用リポジトリ（Streamlit Cloud）"
            "または `~/Documents/DXF-extract-labels/decision_log.csv`（Windows/"
            "ローカル）に追記記録します。同一ラベルについて『ほぼ常に採用されている』"
            "→確定パターン候補、『ほぼ常に非採用』→除外パターン候補として提案します。\n\n"
            "提案はあくまで機械的な集計に基づく候補です。実際にパターンへ反映するかは "
            "`utils/ref_designator.py` の `CONFIRMED_PATTERN_CATEGORIES` /"
            "`EXCLUSION_*_CATEGORIES` を人手で判断・編集してください。"
        )

    st.subheader("入力")
    col_left, col_right = st.columns(2)
    with col_left:
        uploaded_logs = st.file_uploader(
            "decision_log.csv をアップロード（複数可）",
            type="csv",
            accept_multiple_files=True,
            key="decision_log_uploader",
        )
    with col_right:
        folder_text = st.text_area(
            "フォルダパス（1行に1つ、ローカル/Dropbox のフルパス）",
            value="",
            help="ローカル・Windows アプリの decision_log.csv を配置しているフォルダを"
                 "指定すると、配下のログファイルも対象に加えます。",
            key="decision_log_folder_text",
        )
        glob_pattern = st.text_input(
            "検索パターン（glob）", value="decision_log*.csv", key="decision_log_glob")
        recursive = st.checkbox(
            "サブフォルダも検索する", value=True, key="decision_log_recursive")

    with st.expander("☁️ GitHub のログ専用リポジトリから直接取得（Streamlit Cloud）", expanded=False):
        st.caption(
            "アプリ本体と同じ `st.secrets['github']` を使う場合はトークン欄を空にしてください。"
        )
        gh_repo = st.text_input(
            "リポジトリ（owner/repo）", value="", key="decision_log_gh_repo")
        gh_path = st.text_input(
            "パス", value=dlog.DEFAULT_GITHUB_PATH, key="decision_log_gh_path")
        gh_branch = st.text_input(
            "ブランチ", value=dlog.DEFAULT_GITHUB_BRANCH, key="decision_log_gh_branch")
        gh_token = st.text_input(
            "トークン（省略時は st.secrets['github']['token'] を使用）",
            value="", type="password", key="decision_log_gh_token")
        fetch_from_github = st.checkbox(
            "この設定で GitHub から取得する", value=False, key="decision_log_gh_enable")

    folder_paths = [line for line in folder_text.splitlines() if line.strip()]

    col1, col2, col3 = st.columns(3)
    min_occurrences = col1.number_input(
        "最小出現回数（未満は様子見）", min_value=1, value=3, step=1,
        key="decision_log_min_occurrences")
    confirm_rate = col2.slider(
        "確定パターン候補とする採用率の下限", min_value=0.5, max_value=1.0,
        value=1.0, step=0.05, key="decision_log_confirm_rate")
    exclude_rate = col3.slider(
        "除外パターン候補とする非採用率の下限", min_value=0.5, max_value=1.0,
        value=1.0, step=0.05, key="decision_log_exclude_rate")

    if st.button("判断ログを集計する", type="primary"):
        sources = _iter_input_sources(uploaded_logs, folder_paths, glob_pattern, recursive)

        if fetch_from_github:
            if not gh_repo:
                st.warning("GitHub から取得するにはリポジトリ（owner/repo）を指定してください。")
            else:
                token = gh_token or None
                if not token:
                    try:
                        token = st.secrets.get('github', {}).get('token')
                    except Exception:
                        token = None
                if not token:
                    st.warning("トークンが指定されておらず、st.secrets['github']['token'] も"
                               "見つかりません。")
                else:
                    try:
                        with st.spinner(f"GitHub（{gh_repo}）から取得中..."):
                            text = dlog.fetch_log_text(
                                token=token, repo=gh_repo,
                                path=gh_path or dlog.DEFAULT_GITHUB_PATH,
                                branch=gh_branch or dlog.DEFAULT_GITHUB_BRANCH,
                            )
                        if text:
                            sources.append((f"GitHub:{gh_repo}", text))
                        else:
                            st.info(f"GitHub（{gh_repo}）にログファイルがまだありません。")
                    except Exception as e:
                        st.warning(f"GitHub からの取得に失敗しました: {e}")

        if not sources:
            st.warning("入力ログがありません。アップロード・フォルダ指定・GitHub取得の"
                       "いずれかを行ってください。")
        else:
            with st.spinner(f"{len(sources)} 個のログソースを処理中..."):
                agg, per_source_stats = aggregate_decision_log(sources)
                suggestion_rows = build_decision_log_suggestions(
                    agg, min_occurrences=min_occurrences,
                    confirm_rate=confirm_rate, exclude_rate=exclude_rate)
                excel_bytes = build_decision_log_workbook(suggestion_rows)

            st.session_state['rda_decision_log_result'] = {
                'excel_bytes': excel_bytes,
                'sources': [name for name, _ in sources],
                'per_source_stats': per_source_stats,
                'rows': suggestion_rows,
                'total_labels': len(agg),
                'confirm_candidates': sum(
                    1 for r in suggestion_rows if r['提案'] == '確定パターン候補'),
                'exclude_candidates': sum(
                    1 for r in suggestion_rows if r['提案'] == '除外パターン候補'),
            }
            st.rerun()

    result = st.session_state.get('rda_decision_log_result')
    if result:
        st.subheader("結果")
        st.success(f"{len(result['sources'])} 個のログソースを処理しました。")

        m1, m2, m3 = st.columns(3)
        m1.metric("集計対象ラベル数", result['total_labels'])
        m2.metric("確定パターン候補", result['confirm_candidates'])
        m3.metric("除外パターン候補", result['exclude_candidates'])

        with st.expander("📊 入力ソース別の内訳", expanded=False):
            import pandas as pd
            st.dataframe(pd.DataFrame(result['per_source_stats']), width='stretch', hide_index=True)

        with st.expander("📊 提案一覧", expanded=True):
            import pandas as pd
            st.dataframe(pd.DataFrame(result['rows']), width='stretch', hide_index=True)

        st.download_button(
            label="Excelをダウンロード（DecisionLogSummary）",
            data=result['excel_bytes'],
            file_name="decision_log_summary.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            key="decision_log_download",
        )

        if st.button("🔄 リセット", key="decision_log_reset"):
            del st.session_state['rda_decision_log_result']
            st.rerun()


def app():
    st.title("Reference Designator 抽出検討ツール")
    tab1, tab2 = st.tabs(["extracted_labels 集計", "判断ログ分析（v1.7.0）"])
    with tab1:
        _app_extracted_labels()
    with tab2:
        _app_decision_log()


if __name__ == "__main__":
    app()
